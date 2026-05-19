import importlib.util
import csv
import json
import sys
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory


ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = ROOT / "scripts" / "export_suite_results.py"


def _load_script_module():
    module_name = "suite_results_export_test_module"
    spec = importlib.util.spec_from_file_location(module_name, SCRIPT_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to load script module from {SCRIPT_PATH}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


suite_export = _load_script_module()


class SuiteResultsExportTest(unittest.TestCase):
    def _event(self, role: str, content: str, *, actor: str) -> dict:
        return {
            "view": ["target", "combined"],
            "actor": actor,
            "edit": {
                "type": "add_message",
                "message": {"role": role, "content": content},
            },
        }

    def _score_row(
        self,
        *,
        test_case_id: str,
        behavior: str,
        permissible: bool,
        policy_violation: bool,
        overrefusal: bool,
        node0_relevant: bool,
        node0_violated: bool | None,
        node1_relevant: bool,
        node1_violated: bool | None,
    ) -> dict:
        return {
            "type": "scenario",
            "test_case_id": test_case_id,
            "behavior": "behavior",
            "dimensions": {"behavior": behavior},
            "target": "target-model",
            "tester_model": "tester-model",
            "judge_model": "judge-model",
            "judge_status": "ok",
            "judge_error": None,
            "verdict": {
                "dimensions": {
                    "policy_violation": policy_violation,
                    "overrefusal": overrefusal,
                },
                "justification": "overall",
                "dimension_justifications": {
                    "policy_violation": "taxonomy rationale",
                    "overrefusal": "overrefusal rationale",
                },
                "node_judgments": [
                    {
                        "node_index": 0,
                        "node_name": "node-a",
                        "relevant": node0_relevant,
                        "violated": node0_violated,
                        "confidence": "high" if node0_relevant else "low",
                        "evidence_turns": [2] if node0_relevant else [],
                        "reasoning": "node-a" if node0_relevant else "",
                    },
                    {
                        "node_index": 1,
                        "node_name": "node-b",
                        "relevant": node1_relevant,
                        "violated": node1_violated,
                        "confidence": "high" if node1_relevant else "low",
                        "evidence_turns": [2] if node1_relevant else [],
                        "reasoning": "node-b" if node1_relevant else "",
                    },
                ],
            },
        }

    def _write_fixture_suite(
        self,
        root: Path,
        *,
        old_metrics: bool = False,
        run_b_started_at: str | None = "2026-04-02T02:00:00Z",
    ) -> Path:
        results_root = root / "results"
        suite_dir = results_root / "suite-a"
        run_a_dir = suite_dir / "run-a"
        run_b_dir = suite_dir / "run-b"
        run_a_dir.mkdir(parents=True, exist_ok=True)
        run_b_dir.mkdir(parents=True, exist_ok=True)

        (suite_dir / "suite.json").write_text(
            json.dumps({"created_at": "2026-04-02T00:00:00Z"}),
            encoding="utf-8",
        )
        policy_raw = {
            "behavior": {"name": "Risk Name"},
            "behavior_categories": [
                {"name": "node-a", "definition": "def a", "permissible": False},
                {"name": "node-b", "definition": "def b", "permissible": True},
            ],
        }
        (suite_dir / "taxonomy.json").write_text(json.dumps(policy_raw), encoding="utf-8")
        (suite_dir / "test_set.jsonl").write_text(
            "\n".join(
                [
                    json.dumps(
                        {
                            "type": "scenario",
                            "test_case_id": "test-case-1",
                            "behavior": "behavior",
                            "dimensions": {"behavior": "node-a"},
                            "seed": {
                                "title": "Title A",
                                "description": "Description A",
                                "system_prompt": "Prompt A",
                            },
                        }
                    ),
                    json.dumps(
                        {
                            "type": "scenario",
                            "test_case_id": "test-case-2",
                            "behavior": "behavior",
                            "dimensions": {"behavior": "node-b"},
                            "seed": {
                                "title": "Title B",
                                "description": "Description B",
                                "system_prompt": "Prompt B",
                            },
                        }
                    ),
                ]
            )
            + "\n",
            encoding="utf-8",
        )

        (run_a_dir / "manifest.json").write_text(
            json.dumps({"status": "completed", "started_at": "2026-04-02T01:00:00Z", "ended_at": "2026-04-02T01:05:00Z"}),
            encoding="utf-8",
        )
        (run_b_dir / "manifest.json").write_text(
            json.dumps(
                {
                    "status": "running",
                    **(
                        {"started_at": run_b_started_at}
                        if run_b_started_at is not None
                        else {}
                    ),
                }
            ),
            encoding="utf-8",
        )

        transcript_rows = [
            {
                "type": "scenario",
                "test_case_id": "test-case-1",
                "behavior": "behavior",
                "dimensions": {"behavior": "node-a"},
                "target": "target-model",
                "tester_model": "tester-model",
                "stop_reason": "tester_ended",
                "events": [
                    self._event("user", "Need advice", actor="tester"),
                    self._event("assistant", "Here is a reply", actor="target"),
                ],
            }
        ]
        (run_a_dir / "transcripts.jsonl").write_text(
            "\n".join(json.dumps(row) for row in transcript_rows) + "\n",
            encoding="utf-8",
        )
        (run_b_dir / "transcripts.jsonl").write_text(
            json.dumps(
                {
                    "type": "scenario",
                    "test_case_id": "test-case-2",
                    "behavior": "behavior",
                    "dimensions": {"behavior": "node-b"},
                    "target": "target-model-2",
                    "tester_model": "tester-model-2",
                    "stop_reason": "max_turns",
                    "events": [
                        self._event("user", "Hello", actor="tester"),
                        self._event("assistant", "Hi", actor="target"),
                    ],
                }
            )
            + "\n",
            encoding="utf-8",
        )

        score_rows = [
            self._score_row(
                test_case_id="test-case-1",
                behavior="node-a",
                permissible=False,
                policy_violation=True,
                overrefusal=False,
                node0_relevant=True,
                node0_violated=True,
                node1_relevant=False,
                node1_violated=None,
            ),
            self._score_row(
                test_case_id="test-case-2",
                behavior="node-b",
                permissible=True,
                policy_violation=False,
                overrefusal=True,
                node0_relevant=False,
                node0_violated=None,
                node1_relevant=True,
                node1_violated=False,
            ),
        ]
        (run_a_dir / "scores.jsonl").write_text(
            "\n".join(json.dumps(row) for row in score_rows) + "\n",
            encoding="utf-8",
        )

        if old_metrics:
            (run_a_dir / "metrics.json").write_text(
                json.dumps({"overall_permissible": {}, "overall_not_permissible": {}, "by_behavior": []}),
                encoding="utf-8",
            )

        scenario_sampling_dir = suite_dir / "scenario_sampling"
        scenario_sampling_dir.mkdir(parents=True, exist_ok=True)
        (scenario_sampling_dir / "notes.txt").write_text("ignore me", encoding="utf-8")

        return results_root

    def test_csv_and_html_exports_consolidate_one_suite(self) -> None:
        with TemporaryDirectory() as tmp_dir:
            results_root = self._write_fixture_suite(Path(tmp_dir))

            rc = suite_export.main(
                [
                    "--results-root",
                    str(results_root),
                    "--suite",
                    "suite-a",
                    "--format",
                    "csv",
                    "--format",
                    "html",
                ]
            )

            self.assertEqual(rc, 0)
            export_dir = results_root / "suite-a" / "exports"
            self.assertTrue((export_dir / "runs.csv").exists())
            self.assertTrue((export_dir / "test_set.csv").exists())
            self.assertTrue((export_dir / "conversations.csv").exists())
            self.assertTrue((export_dir / "scores.csv").exists())
            self.assertTrue((export_dir / "relevant_nodes.csv").exists())
            self.assertTrue((export_dir / "suite_results.html").exists())

            runs_lines = (export_dir / "runs.csv").read_text(encoding="utf-8").splitlines()
            self.assertEqual(len(runs_lines), 3)
            self.assertIn("suite_id,run_id,status", runs_lines[0])

            test_set_lines = (export_dir / "test_set.csv").read_text(encoding="utf-8").splitlines()
            self.assertEqual(len(test_set_lines), 3)
            self.assertTrue(test_set_lines[0].startswith("suite_id,test_case_id,type"))
            self.assertNotIn("run_id", test_set_lines[0])

            conversations_text = (export_dir / "conversations.csv").read_text(encoding="utf-8")
            self.assertIn("Here is a reply", conversations_text)

            scores_text = (export_dir / "scores.csv").read_text(encoding="utf-8")
            self.assertIn("policy_violation", scores_text)
            self.assertIn("overrefusal_justification", scores_text)
            self.assertIn("node-a_relevant", scores_text)
            self.assertIn("node-a_violated", scores_text)
            self.assertIn("node-a_confidence", scores_text)
            self.assertIn("node-a_reasoning", scores_text)
            self.assertIn("node-b_relevant", scores_text)
            self.assertIn("node-b_violated", scores_text)

            scores_lines = scores_text.splitlines()
            scores_reader = csv.DictReader(scores_lines)
            score_dicts = list(scores_reader)
            seed1_row = next(r for r in score_dicts if r["test_case_id"] == "test-case-1")
            self.assertEqual(seed1_row["node-a_relevant"], "True")
            self.assertEqual(seed1_row["node-a_violated"], "True")
            self.assertEqual(seed1_row["node-a_confidence"], "high")
            self.assertEqual(seed1_row["node-b_relevant"], "False")
            self.assertEqual(seed1_row["node-b_violated"], "")

            relevant_text = (export_dir / "relevant_nodes.csv").read_text(encoding="utf-8")
            self.assertIn("node_violated_rate", relevant_text)
            self.assertIn("policy_violation_rate", relevant_text)

            html_text = (export_dir / "suite_results.html").read_text(encoding="utf-8")
            self.assertIn("Suite Results", html_text)
            self.assertIn("suite-a", html_text)
            self.assertIn("Relevant Nodes", html_text)
            self.assertIn("Description A", html_text)

    def test_excel_export_writes_expected_sheets(self) -> None:
        if importlib.util.find_spec("openpyxl") is None:
            self.skipTest("openpyxl not installed")

        from openpyxl import load_workbook

        with TemporaryDirectory() as tmp_dir:
            results_root = self._write_fixture_suite(Path(tmp_dir))

            rc = suite_export.main(
                [
                    "--results-root",
                    str(results_root),
                    "--suite",
                    "suite-a",
                    "--format",
                    "excel",
                ]
            )

            self.assertEqual(rc, 0)
            workbook_path = results_root / "suite-a" / "exports" / "suite_results.xlsx"
            self.assertTrue(workbook_path.exists())
            workbook = load_workbook(workbook_path)
            self.assertEqual(
                workbook.sheetnames,
                ["Runs", "Test Set", "Conversations", "Scores", "Relevant Nodes"],
            )
            runs_sheet = workbook["Runs"]
            self.assertEqual(runs_sheet["A1"].value, "suite_id")
            self.assertEqual(runs_sheet["B2"].value, "run-a")

    def test_default_export_writes_csv_only(self) -> None:
        with TemporaryDirectory() as tmp_dir:
            results_root = self._write_fixture_suite(Path(tmp_dir))

            rc = suite_export.main(
                [
                    "--results-root",
                    str(results_root),
                    "--suite",
                    "suite-a",
                ]
            )

            self.assertEqual(rc, 0)
            export_dir = results_root / "suite-a" / "exports"
            self.assertTrue((export_dir / "runs.csv").exists())
            self.assertFalse((export_dir / "suite_results.xlsx").exists())
            self.assertFalse((export_dir / "suite_results.html").exists())

    def test_export_can_rerun_after_exports_exist(self) -> None:
        with TemporaryDirectory() as tmp_dir:
            results_root = self._write_fixture_suite(Path(tmp_dir))

            first_rc = suite_export.main(
                [
                    "--results-root",
                    str(results_root),
                    "--suite",
                    "suite-a",
                    "--format",
                    "csv",
                ]
            )
            second_rc = suite_export.main(
                [
                    "--results-root",
                    str(results_root),
                    "--suite",
                    "suite-a",
                    "--format",
                    "html",
                ]
            )

            self.assertEqual(first_rc, 0)
            self.assertEqual(second_rc, 0)
            export_dir = results_root / "suite-a" / "exports"
            self.assertTrue((export_dir / "runs.csv").exists())
            self.assertTrue((export_dir / "suite_results.html").exists())

    def test_missing_run_started_at_stays_blank(self) -> None:
        with TemporaryDirectory() as tmp_dir:
            results_root = self._write_fixture_suite(
                Path(tmp_dir),
                run_b_started_at=None,
            )

            rc = suite_export.main(
                [
                    "--results-root",
                    str(results_root),
                    "--suite",
                    "suite-a",
                    "--format",
                    "csv",
                ]
            )

            self.assertEqual(rc, 0)
            runs_lines = (
                results_root / "suite-a" / "exports" / "runs.csv"
            )
            with runs_lines.open(encoding="utf-8", newline="") as handle:
                rows = list(csv.DictReader(handle))
            self.assertEqual(rows[1]["run_id"], "run-b")
            self.assertEqual(rows[1]["started_at"], "")

    def test_export_ignores_stale_metrics_file(self) -> None:
        with TemporaryDirectory() as tmp_dir:
            results_root = self._write_fixture_suite(Path(tmp_dir), old_metrics=True)

            rc = suite_export.main(
                [
                    "--results-root",
                    str(results_root),
                    "--suite",
                    "suite-a",
                    "--format",
                    "csv",
                ]
            )

            self.assertEqual(rc, 0)
            relevant_text = (results_root / "suite-a" / "exports" / "relevant_nodes.csv").read_text(encoding="utf-8")
            self.assertIn("node-a", relevant_text)
            self.assertIn("1.0", relevant_text)

    def test_build_score_metrics_excludes_failed_judgments_from_event_rates(self) -> None:
        rows = [
            {
                "test_case_id": "ok",
                "behavior": "behavior",
                "dimensions": {"behavior": "behavior"},
                "judge_model": "judge",
                "judge_status": "ok",
                "judge_error": None,
                "verdict": {
                    "dimensions": {
                        "policy_violation": False,
                        "overrefusal": False,
                    },
                    "justification": "good",
                    "node_judgments": [],
                },
            },
            {
                "test_case_id": "bad",
                "behavior": "behavior",
                "dimensions": {"behavior": "behavior"},
                "judge_model": "judge",
                "judge_status": "judge_failed",
                "judge_error": "judge_failed",
                "verdict": {"error": "judge_failed"},
            },
        ]

        metrics = suite_export._build_score_metrics(
            rows,
            policy_raw={
                "behavior": {"name": "behavior"},
                "behavior_categories": [{"name": "behavior", "definition": "def", "permissible": True}],
            },
        )
        overall = metrics["overall_permissible"]
        [node_metrics] = metrics["by_relevant_node"]

        self.assertEqual(overall["count"], 2)
        self.assertEqual(overall["scored_count"], 1)
        self.assertEqual(overall["judge_failures"], 1)
        self.assertEqual(overall["judge_failure_rate"], 0.5)
        self.assertEqual(overall["policy_violation_rate"], 0.0)
        self.assertEqual(overall["overrefusal_rate"], 0.0)
        self.assertEqual(overall["dimensions"]["policy_violation"]["rate"], 0.0)
        self.assertEqual(overall["dimensions"]["policy_violation"]["clear_count"], 1)
        self.assertEqual(overall["dimensions"]["overrefusal"]["rate"], 0.0)
        self.assertEqual(node_metrics["support"], 0)
        self.assertEqual(node_metrics["node_violated_count"], 0)
        self.assertEqual(node_metrics["node_violated_rate"], 0.0)
        self.assertEqual(node_metrics["dimensions"]["policy_violation"]["count"], 0)
        self.assertEqual(node_metrics["dimensions"]["policy_violation"]["rate"], 0.0)

    def test_build_score_metrics_adds_per_node_conditional_rates(self) -> None:
        rows = [
            {
                "test_case_id": "row-1",
                "behavior": "behavior",
                "dimensions": {"behavior": "node-a"},
                "judge_model": "judge",
                "judge_status": "ok",
                "judge_error": None,
                "verdict": {
                    "dimensions": {
                        "policy_violation": True,
                        "overrefusal": False,
                    },
                    "justification": "bad",
                    "node_judgments": [
                        {
                            "node_index": 0,
                            "node_name": "node-a",
                            "relevant": True,
                            "violated": True,
                            "confidence": "high",
                            "evidence_turns": [2],
                            "reasoning": "bad",
                        },
                        {
                            "node_index": 1,
                            "node_name": "node-b",
                            "relevant": False,
                            "violated": None,
                            "confidence": "low",
                            "evidence_turns": [],
                            "reasoning": "",
                        },
                        {
                            "node_index": 2,
                            "node_name": "node-c",
                            "relevant": False,
                            "violated": None,
                            "confidence": "low",
                            "evidence_turns": [],
                            "reasoning": "",
                        },
                    ],
                },
            },
            {
                "test_case_id": "row-2",
                "behavior": "behavior",
                "dimensions": {"behavior": "node-b"},
                "judge_model": "judge",
                "judge_status": "ok",
                "judge_error": None,
                "verdict": {
                    "dimensions": {
                        "policy_violation": False,
                        "overrefusal": True,
                    },
                    "justification": "mixed",
                    "node_judgments": [
                        {
                            "node_index": 0,
                            "node_name": "node-a",
                            "relevant": True,
                            "violated": False,
                            "confidence": "medium",
                            "evidence_turns": [3],
                            "reasoning": "mixed",
                        },
                        {
                            "node_index": 1,
                            "node_name": "node-b",
                            "relevant": True,
                            "violated": False,
                            "confidence": "high",
                            "evidence_turns": [3],
                            "reasoning": "mixed",
                        },
                        {
                            "node_index": 2,
                            "node_name": "node-c",
                            "relevant": False,
                            "violated": None,
                            "confidence": "low",
                            "evidence_turns": [],
                            "reasoning": "",
                        },
                    ],
                },
            },
            {
                "test_case_id": "row-3",
                "behavior": "behavior",
                "dimensions": {"behavior": "node-c"},
                "judge_model": "judge",
                "judge_status": "judge_failed",
                "judge_error": "judge_failed",
                "verdict": {"error": "judge_failed"},
            },
        ]

        metrics = suite_export._build_score_metrics(
            rows,
            policy_raw={
                "behavior": {"name": "behavior"},
                "behavior_categories": [
                    {"name": "node-a", "definition": "def a", "permissible": False},
                    {"name": "node-b", "definition": "def b", "permissible": True},
                    {"name": "node-c", "definition": "def c", "permissible": False},
                ],
            },
        )
        node_a, node_b, node_c = metrics["by_relevant_node"]

        self.assertEqual(node_a["node_index"], 0)
        self.assertEqual(node_a["node_name"], "node-a")
        self.assertFalse(node_a["permissible"])
        self.assertEqual(node_a["support"], 2)
        self.assertEqual(node_a["node_violated_count"], 1)
        self.assertEqual(node_a["node_violated_rate"], 0.5)
        self.assertEqual(node_a["dimensions"]["policy_violation"]["count"], 2)
        self.assertEqual(node_a["dimensions"]["policy_violation"]["flagged_count"], 1)
        self.assertEqual(node_a["dimensions"]["policy_violation"]["clear_count"], 1)
        self.assertEqual(node_a["dimensions"]["policy_violation"]["rate"], 0.5)
        self.assertEqual(node_a["dimensions"]["overrefusal"]["rate"], 0.5)

        self.assertEqual(node_b["node_index"], 1)
        self.assertEqual(node_b["support"], 1)
        self.assertEqual(node_b["node_violated_count"], 0)
        self.assertEqual(node_b["node_violated_rate"], 0.0)
        self.assertEqual(node_b["dimensions"]["policy_violation"]["count"], 1)
        self.assertEqual(node_b["dimensions"]["policy_violation"]["rate"], 0.0)
        self.assertEqual(node_b["dimensions"]["overrefusal"]["rate"], 1.0)

        self.assertEqual(node_c["node_index"], 2)
        self.assertEqual(node_c["node_name"], "node-c")
        self.assertEqual(node_c["support"], 0)
        self.assertEqual(node_c["node_violated_count"], 0)
        self.assertEqual(node_c["node_violated_rate"], 0.0)
        self.assertEqual(node_c["dimensions"]["policy_violation"]["count"], 0)
        self.assertEqual(node_c["dimensions"]["policy_violation"]["flagged_count"], 0)
        self.assertEqual(node_c["dimensions"]["policy_violation"]["clear_count"], 0)
        self.assertEqual(node_c["dimensions"]["policy_violation"]["rate"], 0.0)


class FormatHtmlValueTest(unittest.TestCase):
    def test_bool_before_int(self) -> None:
        result = suite_export._format_html_value(True, "permissible")
        self.assertIn("badge-yes", result)
        result = suite_export._format_html_value(False, "permissible")
        self.assertIn("badge-no", result)

    def test_int_rate_gets_rate_badge(self) -> None:
        result = suite_export._format_html_value(0, "policy_violation_rate")
        self.assertIn("rate-badge", result)
        self.assertIn("0.0%", result)

    def test_float_rate_gets_rate_badge(self) -> None:
        result = suite_export._format_html_value(0.72, "policy_violation_rate")
        self.assertIn("rate-badge", result)
        self.assertIn("72.0%", result)

    def test_nan_and_inf_rates_show_emdash(self) -> None:
        for val in (float("nan"), float("inf"), float("-inf")):
            result = suite_export._format_html_value(val, "some_rate")
            self.assertIn("empty", result)
            self.assertNotIn("rate-badge", result)

    def test_none_and_empty_show_emdash(self) -> None:
        for val in (None, ""):
            result = suite_export._format_html_value(val, "any_header")
            self.assertIn("empty", result)

    def test_int_zero_not_treated_as_empty(self) -> None:
        result = suite_export._format_html_value(0, "count")
        self.assertEqual(result, "0")

    def test_xss_escaped(self) -> None:
        result = suite_export._format_html_value('<script>alert(1)</script>', "name")
        self.assertNotIn("<script>", result)
        self.assertIn("&lt;script&gt;", result)


if __name__ == "__main__":
    unittest.main()
