# Copyright (c) Microsoft Corporation.
# Licensed under the MIT License.

"""Export one suite's results as csv tables, an Excel workbook, or html."""

from __future__ import annotations

import argparse
import csv
import html
import math
import re
from pathlib import Path
from typing import Any

from assert_ai.core.io import (
    get_permissible_flag,
    load_json,
    load_jsonl,
    definitions_by_behavior,
    permissible_by_behavior,
    policy_definition,
    policy_permissible,
    resolve_path,
    row_behavior,
    row_factors,
)
from assert_ai.core.judge import get_verdict_dimension, infer_judge_status, is_valid_event_flag
from assert_ai.core.transcript import Transcript, TranscriptEvent, TranscriptMetadata

EXPORT_DIR_NAME = "exports"
CSV_FORMAT = "csv"
EXCEL_FORMAT = "excel"
HTML_FORMAT = "html"
SUPPORTED_FORMATS = (CSV_FORMAT, EXCEL_FORMAT, HTML_FORMAT)

RUNS_TABLE = "runs"
TEST_SET_TABLE = "test_set"
CONVERSATIONS_TABLE = "conversations"
SCORES_TABLE = "scores"
RELEVANT_NODES_TABLE = "relevant_nodes"

TABLE_ORDER = (
    RUNS_TABLE,
    TEST_SET_TABLE,
    CONVERSATIONS_TABLE,
    SCORES_TABLE,
    RELEVANT_NODES_TABLE,
)

SHEET_NAMES = {
    RUNS_TABLE: "Runs",
    TEST_SET_TABLE: "Test Set",
    CONVERSATIONS_TABLE: "Conversations",
    SCORES_TABLE: "Scores",
    RELEVANT_NODES_TABLE: "Relevant Nodes",
}

CSV_FILENAMES = {
    RUNS_TABLE: "runs.csv",
    TEST_SET_TABLE: "test_set.csv",
    CONVERSATIONS_TABLE: "conversations.csv",
    SCORES_TABLE: "scores.csv",
    RELEVANT_NODES_TABLE: "relevant_nodes.csv",
}


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export one suite's results as csv tables, an Excel workbook, or html.",
    )
    parser.add_argument(
        "--suite",
        required=True,
        help="Suite id under artifacts/results/.",
    )
    parser.add_argument(
        "--results-root",
        default="artifacts/results",
        help="Results root containing suite directories.",
    )
    parser.add_argument(
        "--format",
        action="append",
        choices=SUPPORTED_FORMATS,
        dest="formats",
        help="Export format. Repeat to write multiple formats. Defaults to csv.",
    )
    return parser.parse_args(argv)


def _suite_dir(results_root: str, suite_id: str) -> Path:
    root = resolve_path(results_root)
    suite_dir = root / suite_id
    if not suite_dir.is_dir():
        raise ValueError(f"Suite not found: {suite_dir}")
    return suite_dir


def _run_dirs(suite_dir: Path) -> list[Path]:
    return sorted(
        path
        for path in suite_dir.iterdir()
        if path.is_dir() and (path / "manifest.json").is_file()
    )


def _first_nonempty(rows: list[dict[str, Any]], key: str) -> str:
    for row in rows:
        value = row.get(key)
        if isinstance(value, str) and value:
            return value
    return ""


def _event_rate(rows: list[dict[str, Any]], dimension: str) -> float | None:
    scored_rows = [row for row in rows if infer_judge_status(row) == "ok"]
    values: list[bool] = []
    for row in scored_rows:
        value = get_verdict_dimension(row.get("verdict"), dimension)
        if is_valid_event_flag(value):
            values.append(bool(value))
    if not values:
        return None
    return sum(1 for value in values if value) / len(values)


def _judge_failure_count(rows: list[dict[str, Any]]) -> int:
    return sum(1 for row in rows if infer_judge_status(row) != "ok")


def _row_permissible(
    row: dict[str, Any],
    permissible_by_name: dict[str, bool],
) -> bool:
    return policy_permissible(permissible_by_name, row_behavior(row))


def _test_case_definition(
    test_case_row: dict[str, Any],
    definitions_by_name: dict[str, str],
) -> str:
    return policy_definition(definitions_by_name, row_behavior(test_case_row))


def _build_score_metrics(records: list[dict[str, Any]], *, policy_raw: dict[str, Any]) -> dict[str, Any]:
    behavior_categories = policy_raw.get("behavior_categories")
    if not isinstance(behavior_categories, list):
        raise ValueError("taxonomy.json must contain a behavior_categories list")

    permissible_by_name = permissible_by_behavior(policy_raw)
    permissible_records = [
        row for row in records if _row_permissible(row, permissible_by_name)
    ]
    not_permissible_records = [
        row for row in records if not _row_permissible(row, permissible_by_name)
    ]
    all_dims = sorted(
        {
            key
            for row in records
            for key, value in (((row.get("verdict") or {}).get("dimensions")) or {}).items()
            if isinstance(row.get("verdict"), dict) and is_valid_event_flag(value)
        }
    )

    def compute_stats(items: list[dict[str, Any]]) -> dict[str, Any]:
        scored_items = [item for item in items if infer_judge_status(item) == "ok"]
        total = len(items)
        failures = total - len(scored_items)
        result: dict[str, Any] = {
            "count": total,
            "scored_count": len(scored_items),
            "judge_failures": failures,
            "judge_failure_rate": failures / total if total else 0.0,
            "dimensions": {},
        }
        for dim in all_dims:
            raw_values = []
            for item in scored_items:
                value = get_verdict_dimension(item.get("verdict"), dim)
                if is_valid_event_flag(value):
                    raw_values.append(bool(value))
            flagged_count = sum(1 for value in raw_values if value)
            clear_count = len(raw_values) - flagged_count
            rate = flagged_count / len(raw_values) if raw_values else None
            result["dimensions"][dim] = {
                "count": len(raw_values),
                "flagged_count": flagged_count,
                "clear_count": clear_count,
                "rate": rate,
            }
            result[f"{dim}_rate"] = rate
        return result

    by_behavior: dict[str, dict[str, Any]] = {}
    for row in records:
        behavior = row_behavior(row)
        payload = by_behavior.setdefault(
            behavior,
            {"permissible": _row_permissible(row, permissible_by_name), "items": []},
        )
        payload["items"].append(row)

    scored_records = [row for row in records if infer_judge_status(row) == "ok"]
    by_relevant_node = []
    for node_index, node_payload in enumerate(behavior_categories):
        node_name = str(node_payload.get("name") or "") if isinstance(node_payload, dict) else ""
        relevant_rows: list[dict[str, Any]] = []
        node_violated_count = 0
        for row in scored_records:
            verdict = row.get("verdict")
            if not isinstance(verdict, dict):
                continue
            node_judgments = verdict.get("node_judgments")
            if not isinstance(node_judgments, list) or node_index >= len(node_judgments):
                continue
            node = node_judgments[node_index]
            if not isinstance(node, dict) or node.get("relevant") is not True:
                continue
            relevant_rows.append(row)
            if node.get("violated") is True:
                node_violated_count += 1

        support = len(relevant_rows)
        node_dimensions: dict[str, dict[str, Any]] = {}
        for dim in all_dims:
            flagged_count = 0
            for row in relevant_rows:
                value = get_verdict_dimension(row.get("verdict"), dim)
                if not is_valid_event_flag(value):
                    raise ValueError(f"missing or invalid dimension '{dim}' in scored row")
                flagged_count += int(value)
            clear_count = support - flagged_count
            node_dimensions[dim] = {
                "count": support,
                "flagged_count": flagged_count,
                "clear_count": clear_count,
                "rate": flagged_count / support if support else 0.0,
            }

        by_relevant_node.append(
            {
                "node_index": node_index,
                "node_name": node_name,
                "permissible": (
                    get_permissible_flag(node_payload, default=True)
                    if isinstance(node_payload, dict)
                    else True
                ),
                "support": support,
                "node_violated_count": node_violated_count,
                "node_violated_rate": node_violated_count / support if support else 0.0,
                "dimensions": node_dimensions,
            }
        )

    return {
        "overall_permissible": compute_stats(permissible_records),
        "overall_not_permissible": compute_stats(not_permissible_records),
        "by_behavior": [
            {
                "behavior": behavior,
                "permissible": payload["permissible"],
                **compute_stats(payload["items"]),
            }
            for behavior, payload in sorted(by_behavior.items())
        ],
        "by_relevant_node": by_relevant_node,
    }


def _test_case_counts(test_set_rows: list[dict[str, Any]]) -> tuple[int, int]:
    prompt_count = sum(1 for row in test_set_rows if row.get("type") == "prompt")
    scenario_count = sum(1 for row in test_set_rows if row.get("type") == "scenario")
    return prompt_count, scenario_count


def _transcript_from_row(row: dict[str, Any]) -> Transcript:
    return Transcript(
        metadata=TranscriptMetadata(
            kind=str(row.get("type") or ""),
            test_case_id=str(row.get("test_case_id") or ""),
            behavior=str(row.get("behavior") or ""),
            target=str(row.get("target") or ""),
            tester_model=str(row.get("tester_model") or ""),
            dimensions=row_factors(row),
        ),
        events=[
            TranscriptEvent.model_validate(event)
            for event in row.get("events", [])
            if isinstance(event, dict)
        ],
        stop_reason=str(row.get("stop_reason") or "") or None,
    )


def _test_case_row(
    test_case_row: dict[str, Any],
    suite_id: str,
    permissible_by_name: dict[str, bool],
    definitions_by_name: dict[str, str],
) -> dict[str, Any]:
    payload = test_case_row.get("seed")
    test_case_payload = payload if isinstance(payload, dict) else {}
    kind = str(test_case_row.get("type") or "")
    description = str(test_case_payload.get("description") or "")
    title = str(test_case_payload.get("title") or "") if kind == "scenario" else ""
    return {
        "suite_id": suite_id,
        "test_case_id": str(test_case_row.get("test_case_id") or ""),
        "type": kind,
        "behavior": row_behavior(test_case_row),
        "permissible": _row_permissible(test_case_row, permissible_by_name),
        "definition": _test_case_definition(test_case_row, definitions_by_name),
        "title": title,
        "description": description,
        "system_prompt": str(test_case_payload.get("system_prompt") or ""),
    }


def _score_row(
    score_row: dict[str, Any],
    suite_id: str,
    run_id: str,
    score_dimensions: list[str],
    node_names: list[str],
    permissible_by_name: dict[str, bool],
) -> dict[str, Any]:
    verdict = score_row.get("verdict")
    verdict_payload = verdict if isinstance(verdict, dict) else {}
    dimension_justifications = verdict_payload.get("dimension_justifications")
    justifications = (
        dimension_justifications
        if isinstance(dimension_justifications, dict)
        else {}
    )
    row = {
        "suite_id": suite_id,
        "run_id": run_id,
        "test_case_id": str(score_row.get("test_case_id") or ""),
        "type": str(score_row.get("type") or ""),
        "behavior": row_behavior(score_row),
        "permissible": _row_permissible(score_row, permissible_by_name),
        "judge_status": str(score_row.get("judge_status") or ""),
        "judge_error": str(score_row.get("judge_error") or ""),
        "target": str(score_row.get("target") or ""),
        "tester_model": str(score_row.get("tester_model") or ""),
        "judge_model": str(score_row.get("judge_model") or ""),
        "justification": str(verdict_payload.get("justification") or ""),
    }
    for dimension in score_dimensions:
        value = get_verdict_dimension(verdict_payload, dimension)
        row[dimension] = value if is_valid_event_flag(value) else ""
        justification = justifications.get(dimension)
        row[f"{dimension}_justification"] = (
            str(justification) if isinstance(justification, str) else ""
        )

    node_judgments_by_name = _index_node_judgments(verdict_payload)
    for name in node_names:
        judgment = node_judgments_by_name.get(name)
        if judgment is not None:
            row[f"{name}_relevant"] = judgment.get("relevant")
            row[f"{name}_violated"] = judgment.get("violated")
            row[f"{name}_confidence"] = str(judgment.get("confidence") or "")
            row[f"{name}_reasoning"] = str(judgment.get("reasoning") or "")
        else:
            row[f"{name}_relevant"] = ""
            row[f"{name}_violated"] = ""
            row[f"{name}_confidence"] = ""
            row[f"{name}_reasoning"] = ""
    return row


def _index_node_judgments(verdict: dict[str, Any]) -> dict[str, dict[str, Any]]:
    """Build a dict from node_name → judgment entry for fast lookup."""
    raw = verdict.get("node_judgments")
    if not isinstance(raw, list):
        return {}
    result: dict[str, dict[str, Any]] = {}
    for entry in raw:
        if isinstance(entry, dict):
            name = entry.get("node_name")
            if isinstance(name, str) and name:
                result[name] = entry
    return result


def _relevant_node_row(
    node_row: dict[str, Any],
    suite_id: str,
    run_id: str,
    relevant_dimensions: list[str],
) -> dict[str, Any]:
    dimensions_payload = node_row.get("dimensions")
    dimensions = dimensions_payload if isinstance(dimensions_payload, dict) else {}
    row = {
        "suite_id": suite_id,
        "run_id": run_id,
        "node_index": node_row.get("node_index"),
        "node_name": str(node_row.get("node_name") or ""),
        "node_permissible": node_row.get("permissible"),
        "support": node_row.get("support"),
        "node_violated_count": node_row.get("node_violated_count"),
        "node_violated_rate": node_row.get("node_violated_rate"),
    }
    for dimension in relevant_dimensions:
        summary_payload = dimensions.get(dimension)
        summary = summary_payload if isinstance(summary_payload, dict) else {}
        row[f"{dimension}_count"] = summary.get("count", "")
        row[f"{dimension}_flagged_count"] = summary.get("flagged_count", "")
        row[f"{dimension}_clear_count"] = summary.get("clear_count", "")
        row[f"{dimension}_rate"] = summary.get("rate", "")
    return row


_LONG_TEXT_HEADERS = frozenset(
    {
        "description",
        "system_prompt",
        "conversation_text",
        "justification",
        "definition",
    }
)


def _display_header(name: str) -> str:
    """Convert snake_case field name to a human-readable label."""
    return name.replace("_", " ").title().replace(" Id", " ID")


def _rate_severity(value: float) -> str:
    pct = value * 100
    if pct > 50:
        return "sev-high"
    if pct > 20:
        return "sev-mid"
    return "sev-low"


_TURN_RE = re.compile(
    r"^(System:|\[Turn \d+\] (?:User|Assistant):)", re.MULTILINE
)

_ROLE_MAP = {"system": "system", "user": "user", "assistant": "assistant"}


def _format_conversation(text: str) -> str:
    """Parse a transcript string into styled HTML turn blocks."""
    parts = _TURN_RE.split(text)
    if len(parts) <= 1:
        return f'<div class="conv-body">{html.escape(text)}</div>'
    turns: list[str] = []
    # parts[0] is text before first marker (usually empty)
    if parts[0].strip():
        body = html.escape(parts[0].strip())
        turns.append(
            f'<div class="conv-turn conv-system">'
            f'<div class="conv-body">{body}</div></div>'
        )
    for i in range(1, len(parts), 2):
        marker = parts[i]
        body = html.escape(parts[i + 1].strip()) if i + 1 < len(parts) else ""
        if marker.startswith("System"):
            role = "system"
        elif "User:" in marker:
            role = "user"
        else:
            role = "assistant"
        label = html.escape(marker.rstrip(":"))
        turns.append(
            f'<div class="conv-turn conv-{role}">'
            f'<div class="conv-role conv-role-{role}">{label}</div>'
            f'<div class="conv-body">{body}</div></div>'
        )
    return "".join(turns)


def _format_html_value(value: Any, header: str) -> str:
    if value is None or value == "":
        return '<span class="empty">\u2014</span>'
    if isinstance(value, bool):
        cls = "badge-yes" if value else "badge-no"
        label = "Yes" if value else "No"
        return f'<span class="badge {cls}">{label}</span>'
    if isinstance(value, (int, float)) and not isinstance(value, bool) and header.endswith("_rate"):
        fval = float(value)
        if not math.isfinite(fval):
            return '<span class="empty">\u2014</span>'
        pct = fval * 100
        cls = "rate-high" if pct > 50 else ("rate-mid" if pct > 20 else "rate-low")
        return f'<span class="rate-badge {cls}">{pct:.1f}%</span>'
    text_raw = str(value)
    text = html.escape(text_raw)
    if header in ("status", "judge_status"):
        lower = text_raw.lower()
        if lower in ("completed", "ok"):
            cls = "badge-yes"
        elif lower in ("failed", "error"):
            cls = "badge-no"
        else:
            cls = "badge-neutral"
        return f'<span class="badge {cls}">{text}</span>'
    is_long = header in _LONG_TEXT_HEADERS or header.endswith("_justification") or header.endswith("_reasoning")
    is_conv = header == "conversation_text"
    if is_conv:
        formatted = _format_conversation(text_raw)
        if len(text_raw) > 300:
            preview = html.escape(text_raw[:300])
            return (
                '<div class="expandable">'
                f'<div class="preview">{preview}\u2026</div>'
                f'<div class="full-text" hidden>{formatted}</div>'
                '<button class="expand-btn" aria-expanded="false" '
                'onclick="openModal(this)">Show conversation</button>'
                "</div>"
            )
        return formatted
    if is_long and len(text_raw) > 300:
        preview = html.escape(text_raw[:300])
        return (
            '<div class="expandable">'
            f'<div class="preview">{preview}\u2026</div>'
            f'<div class="full-text long-text" hidden>{text}</div>'
            '<button class="expand-btn" aria-expanded="false" '
            'onclick="openModal(this)">Show more</button>'
            "</div>"
        )
    if is_long:
        return f'<div class="long-text">{text}</div>'
    return text


def load_suite_tables(
    suite_dir: Path,
    suite_id: str,
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, list[str]]]:
    taxonomy = load_json(suite_dir / "taxonomy.json") or {}
    definitions_by_name = definitions_by_behavior(taxonomy)
    permissible_by_name = permissible_by_behavior(taxonomy)
    test_case_rows = load_jsonl(suite_dir / "test_set.jsonl")
    prompt_test_case_count, scenario_test_case_count = _test_case_counts(test_case_rows)
    run_rows: list[dict[str, Any]] = []
    test_set_table = [
        _test_case_row(test_case_row, suite_id, permissible_by_name, definitions_by_name)
        for test_case_row in test_case_rows
    ]
    conversation_rows: list[dict[str, Any]] = []
    raw_score_rows: list[tuple[str, dict[str, Any]]] = []
    raw_relevant_node_rows: list[tuple[str, dict[str, Any]]] = []
    score_dimensions: set[str] = set()
    relevant_dimensions: set[str] = set()
    score_node_names: list[str] = []
    seen_node_names: set[str] = set()

    behavior_payload = taxonomy.get("behavior") if isinstance(taxonomy.get("behavior"), dict) else {}
    behavior_name = str(behavior_payload.get("name") or "")
    if not behavior_name:
        behavior_name = str(test_case_rows[0].get("behavior") or "") if test_case_rows else ""

    for run_dir in _run_dirs(suite_dir):
        run_id = run_dir.name
        manifest = load_json(run_dir / "manifest.json") or {}
        inference_rows = load_jsonl(run_dir / "inference_set.jsonl")
        score_rows = load_jsonl(run_dir / "scores.jsonl")
        metrics = _build_score_metrics(score_rows, policy_raw=taxonomy) if score_rows else None

        target = _first_nonempty(score_rows, "target") or _first_nonempty(inference_rows, "target")
        tester_model = _first_nonempty(score_rows, "tester_model") or _first_nonempty(inference_rows, "tester_model")
        judge_model = _first_nonempty(score_rows, "judge_model")

        for inference_row in inference_rows:
            transcript = _transcript_from_row(inference_row)
            conversation_rows.append(
                {
                    "suite_id": suite_id,
                    "run_id": run_id,
                    "test_case_id": str(inference_row.get("test_case_id") or ""),
                    "type": str(inference_row.get("type") or ""),
                    "behavior": row_behavior(inference_row),
                    "permissible": _row_permissible(
                        inference_row,
                        permissible_by_name,
                    ),
                    "stop_reason": str(inference_row.get("stop_reason") or ""),
                    "turn_count": transcript.count_turns("target", skip_system=True),
                    "target": str(inference_row.get("target") or ""),
                    "tester_model": str(inference_row.get("tester_model") or ""),
                    "conversation_text": transcript.format_transcript(
                        "target",
                        skip_system=False,
                        numbered=True,
                        number_system=False,
                    ),
                }
            )

        for score_row in score_rows:
            raw_score_rows.append((run_id, score_row))
            verdict = score_row.get("verdict")
            dimensions_payload = verdict.get("dimensions") if isinstance(verdict, dict) else None
            if isinstance(dimensions_payload, dict):
                for key, value in dimensions_payload.items():
                    if is_valid_event_flag(value):
                        score_dimensions.add(str(key))
            node_judgments = verdict.get("node_judgments") if isinstance(verdict, dict) else None
            if isinstance(node_judgments, list):
                for entry in node_judgments:
                    if isinstance(entry, dict):
                        name = entry.get("node_name")
                        if isinstance(name, str) and name and name not in seen_node_names:
                            seen_node_names.add(name)
                            score_node_names.append(name)

        if isinstance(metrics, dict):
            for node_row in metrics.get("by_relevant_node", []):
                if not isinstance(node_row, dict):
                    continue
                raw_relevant_node_rows.append((run_id, node_row))
                dimensions_payload = node_row.get("dimensions")
                if not isinstance(dimensions_payload, dict):
                    continue
                for key in dimensions_payload:
                    relevant_dimensions.add(str(key))

        permissible_scores = [
            row for row in score_rows
            if infer_judge_status(row) == "ok"
            and _row_permissible(row, permissible_by_name)
        ]
        not_permissible_scores = [
            row for row in score_rows
            if infer_judge_status(row) == "ok"
            and not _row_permissible(row, permissible_by_name)
        ]
        run_rows.append(
            {
                "suite_id": suite_id,
                "run_id": run_id,
                "status": str(manifest.get("status") or ""),
                "started_at": str(manifest.get("started_at") or ""),
                "ended_at": str(manifest.get("ended_at") or ""),
                "behavior_name": behavior_name,
                "target": target,
                "tester_model": tester_model,
                "judge_model": judge_model,
                "prompt_test_case_count": prompt_test_case_count,
                "scenario_test_case_count": scenario_test_case_count,
                "score_row_count": len(score_rows),
                "judge_failure_count": _judge_failure_count(score_rows),
                "policy_violation_rate": _event_rate(score_rows, "policy_violation"),
                "overrefusal_rate": _event_rate(score_rows, "overrefusal"),
                "permissible_overrefusal_rate": _event_rate(permissible_scores, "overrefusal"),
                "not_permissible_policy_violation_rate": _event_rate(
                    not_permissible_scores,
                    "policy_violation",
                ),
            }
        )

    score_dimension_names = sorted(score_dimensions)
    relevant_dimension_names = sorted(relevant_dimensions)

    score_table = [
        _score_row(
            score_row,
            suite_id,
            run_id,
            score_dimension_names,
            score_node_names,
            permissible_by_name,
        )
        for run_id, score_row in raw_score_rows
    ]
    relevant_node_table = [
        _relevant_node_row(node_row, suite_id, run_id, relevant_dimension_names)
        for run_id, node_row in raw_relevant_node_rows
    ]

    tables = {
        RUNS_TABLE: run_rows,
        TEST_SET_TABLE: test_set_table,
        CONVERSATIONS_TABLE: conversation_rows,
        SCORES_TABLE: score_table,
        RELEVANT_NODES_TABLE: relevant_node_table,
    }

    columns = {
        RUNS_TABLE: [
            "suite_id",
            "run_id",
            "status",
            "started_at",
            "ended_at",
            "behavior_name",
            "target",
            "tester_model",
            "judge_model",
            "prompt_test_case_count",
            "scenario_test_case_count",
            "score_row_count",
            "judge_failure_count",
            "policy_violation_rate",
            "overrefusal_rate",
            "permissible_overrefusal_rate",
            "not_permissible_policy_violation_rate",
        ],
        TEST_SET_TABLE: [
            "suite_id",
            "test_case_id",
            "type",
            "behavior",
            "permissible",
            "definition",
            "title",
            "description",
            "system_prompt",
        ],
        CONVERSATIONS_TABLE: [
            "suite_id",
            "run_id",
            "test_case_id",
            "type",
            "behavior",
            "permissible",
            "stop_reason",
            "turn_count",
            "target",
            "tester_model",
            "conversation_text",
        ],
        SCORES_TABLE: [
            "suite_id",
            "run_id",
            "test_case_id",
            "type",
            "behavior",
            "permissible",
            "judge_status",
            "judge_error",
            "target",
            "tester_model",
            "judge_model",
            *score_dimension_names,
            "justification",
            *[f"{dimension}_justification" for dimension in score_dimension_names],
            *[
                field
                for name in score_node_names
                for field in (
                    f"{name}_relevant",
                    f"{name}_violated",
                    f"{name}_confidence",
                    f"{name}_reasoning",
                )
            ],
        ],
        RELEVANT_NODES_TABLE: [
            "suite_id",
            "run_id",
            "node_index",
            "node_name",
            "node_permissible",
            "support",
            "node_violated_count",
            "node_violated_rate",
            *[
                field
                for dimension in relevant_dimension_names
                for field in (
                    f"{dimension}_count",
                    f"{dimension}_flagged_count",
                    f"{dimension}_clear_count",
                    f"{dimension}_rate",
                )
            ],
        ],
    }
    return tables, columns


def write_csv_exports(
    tables: dict[str, list[dict[str, Any]]],
    columns: dict[str, list[str]],
    out_dir: Path,
) -> None:
    for table_name in TABLE_ORDER:
        path = out_dir / CSV_FILENAMES[table_name]
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns[table_name])
            writer.writeheader()
            for row in tables[table_name]:
                writer.writerow(row)


def write_excel_export(
    tables: dict[str, list[dict[str, Any]]],
    columns: dict[str, list[str]],
    out_path: Path,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "Excel export requires openpyxl. Install it with `python -m pip install openpyxl`."
        ) from exc

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    long_text_headers = {"description", "system_prompt", "conversation_text", "justification"}

    for table_name in TABLE_ORDER:
        sheet = workbook.create_sheet(title=SHEET_NAMES[table_name])
        headers = columns[table_name]
        rows = tables[table_name]
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        sheet.freeze_panes = "A2"
        for index, header in enumerate(headers, start=1):
            column_letter = get_column_letter(index)
            width = 72 if header in long_text_headers or header.endswith("_justification") else min(max(len(header) + 2, 12), 28)
            sheet.column_dimensions[column_letter].width = width
            if header in long_text_headers or header.endswith("_justification"):
                for cell in sheet[column_letter]:
                    cell.alignment = wrap_alignment

    workbook.save(out_path)


_REPORT_CSS = """
:root {
  --bg: #f8fafc; --surface: #fff; --text: #334155; --text-sec: #64748b;
  --heading: #0f172a; --accent: #2563eb; --accent-lt: #dbeafe;
  --border: #e2e8f0; --hover: #f1f5f9; --stripe: #f8fafc;
  --ok: #15803d; --ok-bg: #dcfce7; --err: #b91c1c; --err-bg: #fee2e2;
  --warn: #92400e; --warn-bg: #fef3c7; --radius: 8px;
  --shadow: 0 1px 3px rgba(0,0,0,.08);
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
  background:var(--bg);color:var(--text);line-height:1.5}
.header{background:var(--heading);color:#fff;padding:28px 40px}
.header h1{font-size:22px;font-weight:600}
.header .sub{color:#94a3b8;font-size:13px;font-family:monospace;margin-top:4px}
.main{max-width:1440px;margin:0 auto;padding:24px 40px 60px}
.summary{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));
  gap:14px;margin-bottom:28px}
.card{background:var(--surface);border:1px solid var(--border);
  border-radius:var(--radius);padding:18px 20px;box-shadow:var(--shadow)}
.card.sev-high{border-left:4px solid var(--err)}
.card.sev-mid{border-left:4px solid var(--warn)}
.card.sev-low{border-left:4px solid var(--ok)}
.card .label{font-size:11px;font-weight:600;text-transform:uppercase;
  letter-spacing:.05em;color:var(--text-sec);margin-bottom:6px}
.card .val{font-size:26px;font-weight:700;color:var(--heading);line-height:1.2}
.card .val.sm{font-size:16px;font-weight:600}
.card .detail{font-size:12px;color:var(--text-sec);margin-top:4px}
.tabs{display:flex;gap:2px;border-bottom:2px solid var(--border)}
.tab{padding:10px 18px;font-size:13px;font-weight:500;color:var(--text-sec);
  background:none;border:none;cursor:pointer;border-bottom:2px solid transparent;
  margin-bottom:-2px;transition:all .15s}
.tab:hover{color:var(--text)}
.tab.active{color:var(--accent);border-bottom-color:var(--accent)}
.tab .cnt{font-size:11px;background:var(--hover);color:var(--text-sec);
  padding:1px 7px;border-radius:10px;margin-left:5px}
.tab.active .cnt{background:var(--accent-lt);color:var(--accent)}
.panel{display:none}.panel.active{display:block}
.toolbar{display:flex;align-items:center;gap:14px;padding:14px 0}
.search{padding:7px 12px;border:1px solid var(--border);border-radius:6px;
  font-size:13px;width:260px;outline:none}
.search:focus{border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-lt)}
.row-info{font-size:12px;color:var(--text-sec)}
.tw{background:var(--surface);border:1px solid var(--border);
  border-radius:var(--radius);box-shadow:var(--shadow);
  max-height:80vh;overflow:auto}
table{border-collapse:collapse;width:100%;font-size:13px}
thead th{background:#f1f5f9;padding:9px 12px;text-align:left;font-weight:600;
  font-size:11px;text-transform:uppercase;letter-spacing:.04em;
  color:var(--text-sec);border-bottom:2px solid var(--border);
  position:sticky;top:0;white-space:nowrap;z-index:1;cursor:pointer;user-select:none}
thead th:hover{background:#e2e8f0}
thead th .sort-arrow{margin-left:4px;font-size:10px;opacity:.4}
thead th.sorted .sort-arrow{opacity:1}
tbody td{padding:7px 12px;border-bottom:1px solid #f1f5f9;
  vertical-align:top;max-width:560px;word-break:break-word}
.conv-cell{min-width:480px;max-width:none}
.text-cell{min-width:280px;max-width:640px}
tbody tr:hover{background:#eef2ff}
tbody tr:nth-child(even){background:var(--stripe)}
tbody tr:nth-child(even):hover{background:#eef2ff}
.empty{color:#6b7280}
.badge{display:inline-block;padding:2px 10px;border-radius:4px;
  font-size:12px;font-weight:500}
.badge-yes{background:var(--ok-bg);color:var(--ok)}
.badge-no{background:var(--err-bg);color:var(--err)}
.badge-neutral{background:var(--hover);color:var(--text-sec)}
.rate-badge{display:inline-block;padding:2px 10px;border-radius:4px;
  font-size:13px;font-weight:600;font-variant-numeric:tabular-nums}
.rate-low{background:var(--ok-bg);color:var(--ok)}
.rate-mid{background:var(--warn-bg);color:var(--warn)}
.rate-high{background:var(--err-bg);color:var(--err)}
.long-text{white-space:pre-wrap;word-break:break-word;
  font-size:13px;line-height:1.6}
.expandable .preview{white-space:pre-wrap;word-break:break-word;
  font-size:13px;line-height:1.6}
.expandable .full-text{font-size:13px;line-height:1.6;word-break:break-word}
.expand-btn{display:inline-block;margin-top:6px;padding:2px 10px;font-size:12px;
  color:var(--accent);background:none;border:1px solid var(--accent);
  border-radius:4px;cursor:pointer}
.expand-btn:hover{background:var(--accent-lt)}
.conv-turn{margin-bottom:6px;padding:8px 12px;border-radius:6px}
.conv-system{background:#f5f3ff;border-left:3px solid #7c3aed}
.conv-user{background:#eff6ff;border-left:3px solid #3b82f6}
.conv-assistant{background:#f0fdf4;border-left:3px solid var(--ok)}
.conv-role{font-size:10px;font-weight:700;text-transform:uppercase;
  letter-spacing:.05em;margin-bottom:3px}
.conv-role-system{color:#7c3aed}
.conv-role-user{color:#2563eb}
.conv-role-assistant{color:var(--ok)}
.conv-body{white-space:pre-wrap;word-break:break-word;font-size:13px;line-height:1.6}
.no-match{text-align:center;padding:32px;color:var(--text-sec);
  font-size:13px;display:none}
.seed-link{color:var(--accent);text-decoration:none;cursor:pointer}
.seed-link:hover{text-decoration:underline}
.row-violated{border-left:3px solid var(--err)}
.row-clean{border-left:3px solid var(--ok)}
.kbd-hint{font-size:11px;color:var(--text-sec);margin-left:auto}
.kbd-hint kbd{background:var(--hover);border:1px solid var(--border);
  border-radius:3px;padding:1px 5px;font-size:10px;font-family:inherit}
.modal-overlay{display:none;position:fixed;inset:0;background:rgba(15,23,42,.55);
  z-index:100;overflow-y:auto;padding:40px 24px}
.modal-overlay.open{display:flex;justify-content:center;align-items:flex-start}
.modal-box{background:var(--surface);border-radius:12px;
  box-shadow:0 8px 32px rgba(0,0,0,.18);
  max-width:960px;width:100%;padding:32px 36px;position:relative;margin:auto}
.modal-close{position:absolute;top:14px;right:18px;background:none;border:none;
  font-size:22px;color:var(--text-sec);cursor:pointer;padding:4px 8px;line-height:1}
.modal-close:hover{color:var(--text)}
.modal-body .conv-turn{margin-bottom:10px;padding:12px 18px;border-radius:8px}
.modal-body .conv-body{font-size:14.5px;line-height:1.7}
.modal-body .conv-role{font-size:11px;margin-bottom:5px}
.modal-body .long-text{font-size:14.5px;line-height:1.7}
"""

_REPORT_JS = """
function switchTab(n){
  document.querySelectorAll('.tab').forEach(function(t){
    var isActive=t.getAttribute('data-t')===n;
    t.classList.toggle('active',isActive);
    t.setAttribute('aria-selected',isActive?'true':'false');
  });
  document.querySelectorAll('.panel').forEach(function(p){p.classList.remove('active')});
  document.getElementById('p-'+n).classList.add('active');
}
var _filterTimers={};
function filterTable(n){
  clearTimeout(_filterTimers[n]);
  _filterTimers[n]=setTimeout(function(){_doFilter(n)},150);
}
function _doFilter(n){
  var panel=document.getElementById('p-'+n);
  var q=panel.querySelector('.search').value.toLowerCase();
  var rows=panel.querySelectorAll('tbody tr');
  var vis=0;
  for(var i=0;i<rows.length;i++){
    var ok=!q||rows[i].textContent.toLowerCase().indexOf(q)!==-1;
    rows[i].style.display=ok?'':'none';
    if(ok)vis++;
  }
  panel.querySelector('.row-info').textContent=
    q?vis+' of '+rows.length+' rows':rows.length+' rows';
  panel.querySelector('.no-match').style.display=vis===0?'block':'none';
}
function openModal(btn){
  var c=btn.parentElement,f=c.querySelector('.full-text');
  var modal=document.getElementById('content-modal');
  modal.querySelector('.modal-body').innerHTML=f.innerHTML;
  modal.classList.add('open');document.body.style.overflow='hidden';
}
function closeModal(){
  var modal=document.getElementById('content-modal');
  modal.classList.remove('open');document.body.style.overflow='';
}
var _sortState={};
function sortTable(tableName,colIdx){
  var panel=document.getElementById('p-'+tableName);
  var tbody=panel.querySelector('tbody');
  var rows=Array.prototype.slice.call(tbody.querySelectorAll('tr'));
  var key=tableName+':'+colIdx;
  var asc=_sortState[key]=!_sortState[key];
  var ths=panel.querySelectorAll('thead th');
  ths.forEach(function(th){th.classList.remove('sorted');
    var a=th.querySelector('.sort-arrow');if(a)a.textContent='\\u2195';});
  ths[colIdx].classList.add('sorted');
  var arrow=ths[colIdx].querySelector('.sort-arrow');
  if(arrow)arrow.textContent=asc?'\\u25B2':'\\u25BC';
  rows.sort(function(a,b){
    var ac=a.children[colIdx],bc=b.children[colIdx];
    var av=ac?ac.textContent.replace('%','').trim():'';
    var bv=bc?bc.textContent.replace('%','').trim():'';
    var an=parseFloat(av),bn=parseFloat(bv);
    if(!isNaN(an)&&!isNaN(bn))return asc?an-bn:bn-an;
    return asc?av.localeCompare(bv):bv.localeCompare(av);
  });
  rows.forEach(function(r){tbody.appendChild(r);});
}
function jumpToSeed(el){
  var seed=el.getAttribute('data-seed');
  var target=el.getAttribute('data-target')||'conversations';
  switchTab(target);
  var panel=document.getElementById('p-'+target);
  var input=panel.querySelector('.search');
  input.value=seed;
  _doFilter(target);
}
document.addEventListener('keydown',function(e){
  var modal=document.getElementById('content-modal');
  if(e.key==='Escape'&&modal.classList.contains('open')){closeModal();return;}
  if(e.target.tagName==='INPUT')return;
  var tabs=['runs','test_set','conversations','scores','relevant_nodes'];
  var cur=document.querySelector('.tab.active');
  var idx=cur?tabs.indexOf(cur.getAttribute('data-t')):-1;
  if(e.key==='ArrowRight'){e.preventDefault();switchTab(tabs[(idx+1)%tabs.length]);}
  if(e.key==='ArrowLeft'){e.preventDefault();switchTab(tabs[(idx-1+tabs.length)%tabs.length]);}
  if(e.key==='/'&&!e.ctrlKey&&!e.metaKey){e.preventDefault();
    var s=document.querySelector('.panel.active .search');if(s)s.focus();}
  if(e.key==='Escape'){var s=document.querySelector('.panel.active .search');
    if(s){s.value='';_doFilter(document.querySelector('.tab.active').getAttribute('data-t'));s.blur();}}
});
"""


def write_html_export(
    suite_id: str,
    tables: dict[str, list[dict[str, Any]]],
    columns: dict[str, list[str]],
    out_path: Path,
) -> None:
    esc_id = html.escape(suite_id)
    runs = tables[RUNS_TABLE]

    # --- summary cards ---
    cards: list[str] = []
    if runs:
        behavior = runs[0].get("behavior_name", "")
        if behavior:
            cards.append(
                f'<div class="card"><div class="label">Behavior</div>'
                f'<div class="val sm">{html.escape(str(behavior))}</div></div>'
            )
        targets = sorted({str(r.get("target") or "") for r in runs} - {""})
        if targets:
            cards.append(
                f'<div class="card"><div class="label">Target</div>'
                f'<div class="val sm">{html.escape(", ".join(targets))}</div></div>'
            )
    cards.append(
        f'<div class="card"><div class="label">Test Set</div>'
        f'<div class="val">{len(tables[TEST_SET_TABLE])}</div></div>'
    )
    cards.append(
        f'<div class="card"><div class="label">Conversations</div>'
        f'<div class="val">{len(tables[CONVERSATIONS_TABLE])}</div></div>'
    )
    if runs:
        vr = runs[0].get("policy_violation_rate")
        if vr is not None and isinstance(vr, (int, float)):
            sev = _rate_severity(float(vr))
            cards.append(
                f'<div class="card {sev}"><div class="label">Violation Rate</div>'
                f'<div class="val">{float(vr) * 100:.1f}%</div></div>'
            )
        orr = runs[0].get("overrefusal_rate")
        if orr is not None and isinstance(orr, (int, float)):
            sev = _rate_severity(float(orr))
            cards.append(
                f'<div class="card {sev}"><div class="label">Overrefusal Rate</div>'
                f'<div class="val">{float(orr) * 100:.1f}%</div></div>'
            )
    summary_html = '<div class="summary">' + "".join(cards) + "</div>" if cards else ""

    # --- tabs ---
    tab_parts: list[str] = []
    for table_name in TABLE_ORDER:
        count = len(tables[table_name])
        is_first = table_name == TABLE_ORDER[0]
        active = " active" if is_first else ""
        selected = "true" if is_first else "false"
        tab_parts.append(
            f'<button class="tab{active}" data-t="{table_name}" '
            f'role="tab" aria-selected="{selected}" '
            f'aria-controls="p-{table_name}" '
            f"onclick=\"switchTab('{table_name}')\">"
            f"{html.escape(SHEET_NAMES[table_name])}"
            f'<span class="cnt">{count}</span></button>'
        )
    tabs_html = '<nav class="tabs" role="tablist">' + "".join(tab_parts) + "</nav>"

    # --- table panels ---
    panel_parts: list[str] = []
    for table_name in TABLE_ORDER:
        rows = tables[table_name]
        headers = columns[table_name]
        active = " active" if table_name == TABLE_ORDER[0] else ""
        sheet_name = html.escape(SHEET_NAMES[table_name])
        lines: list[str] = [
            f'<div class="panel{active}" id="p-{table_name}" role="tabpanel">'
        ]
        lines.append(
            f'<div class="toolbar">'
            f'<input class="search" type="search" placeholder="Filter rows\u2026" '
            f'aria-label="Filter {sheet_name} rows" '
            f"oninput=\"filterTable('{table_name}')\">"
            f'<span class="row-info">{len(rows)} rows</span>'
            f'<span class="kbd-hint">'
            f"<kbd>\u2190</kbd><kbd>\u2192</kbd> tabs \u00b7 "
            f"<kbd>/</kbd> search \u00b7 <kbd>Esc</kbd> clear</span>"
            f"</div>"
        )
        lines.append("<div class=\"tw\"><table><thead><tr>")
        for col_idx, header in enumerate(headers):
            display = html.escape(_display_header(header))
            raw = html.escape(header)
            lines.append(
                f'<th scope="col" title="{raw}" '
                f"onclick=\"sortTable('{table_name}',{col_idx})\">"
                f'{display}<span class="sort-arrow">\u2195</span></th>'
            )
        lines.append("</tr></thead><tbody>")
        is_scores = table_name == SCORES_TABLE
        for row in rows:
            row_cls = ""
            if is_scores:
                pv = row.get("policy_violation")
                if pv is True:
                    row_cls = ' class="row-violated"'
                elif pv is False:
                    row_cls = ' class="row-clean"'
            lines.append(f"<tr{row_cls}>")
            for header in headers:
                val = row.get(header, "")
                if header == "test_case_id" and val:
                    esc_val = html.escape(str(val))
                    cell = (
                        f'<a class="seed-link" data-seed="{esc_val}" '
                        f'data-target="conversations" href="#" '
                        f'onclick="jumpToSeed(this);return false;">{esc_val}</a>'
                    )
                else:
                    cell = _format_html_value(val, header)
                td_cls = ""
                if header == "conversation_text":
                    td_cls = ' class="conv-cell"'
                elif header in _LONG_TEXT_HEADERS or header.endswith("_justification") or header.endswith("_reasoning"):
                    td_cls = ' class="text-cell"'
                lines.append(f"<td{td_cls}>{cell}</td>")
            lines.append("</tr>")
        lines.append("</tbody></table></div>")
        lines.append('<div class="no-match">No matching rows</div>')
        lines.append("</div>")
        panel_parts.append("\n".join(lines))
    panels_html = "\n".join(panel_parts)

    modal_html = (
        '<div class="modal-overlay" id="content-modal" '
        'onclick="if(event.target===this)closeModal()">'
        '<div class="modal-box">'
        '<button class="modal-close" onclick="closeModal()" '
        'aria-label="Close">&times;</button>'
        '<div class="modal-body"></div>'
        "</div></div>"
    )
    page = (
        f"<!doctype html>\n"
        f'<html lang="en">\n'
        f"<head>\n"
        f'<meta charset="utf-8">\n'
        f'<meta name="viewport" content="width=device-width, initial-scale=1">\n'
        f"<title>{esc_id} \u2014 suite results</title>\n"
        f"<style>{_REPORT_CSS}</style>\n"
        f"</head>\n"
        f"<body>\n"
        f'<div class="header">\n'
        f"<h1>Suite Results \u2014 {esc_id}</h1>\n"
        f'<div class="sub">artifacts/results/{esc_id}/</div>\n'
        f"</div>\n"
        f'<div class="main">\n'
        f"{summary_html}\n"
        f"{tabs_html}\n"
        f"{panels_html}\n"
        f"</div>\n"
        f"{modal_html}\n"
        f"<script>{_REPORT_JS}</script>\n"
        f"</body>\n"
        f"</html>"
    )
    out_path.write_text(page, encoding="utf-8")


def export_suite(
    *,
    suite_id: str,
    results_root: str,
    formats: list[str],
) -> Path:
    suite_dir = _suite_dir(results_root, suite_id)
    tables, columns = load_suite_tables(suite_dir, suite_id)
    out_dir = suite_dir / EXPORT_DIR_NAME
    out_dir.mkdir(parents=True, exist_ok=True)

    if CSV_FORMAT in formats:
        write_csv_exports(tables, columns, out_dir)
    if EXCEL_FORMAT in formats:
        write_excel_export(tables, columns, out_dir / "suite_results.xlsx")
    if HTML_FORMAT in formats:
        write_html_export(suite_id, tables, columns, out_dir / "suite_results.html")
    return out_dir


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)
    formats = args.formats or [CSV_FORMAT]
    export_suite(
        suite_id=str(args.suite),
        results_root=str(args.results_root),
        formats=formats,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
